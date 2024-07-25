import openpyxl as pyXL
import PySimpleGUI as sg

class exercise:
    def __init__(self, name, targetMuscle):
        self.name = name
        self.target = targetMuscle
    
    def __str__(self) -> str:
        return(f'{self.name}')
    
    def getName(self):
        return(f'{self.name}')
    

layout = [
    [sg.Text('Division:'), sg.Spin(
        ['Push/pull/legs', 'Upper/lower', 'Full Body'], key='-SPLIT-'
    )],
    [sg.Text('Days'), sg.Spin(
        ['3 Days', '4 Days', '5 Days'], key='-DAYS-'
    )],
    [sg.Text('Modality'), sg.Spin(
        ['Calisthenics', 'Weightlifting', 'Both'], key='-MODALITY-'
    )],
    [sg.Button('Generate', key='-GENERATE-')]
]

window = sg.Window('Planejamento de Treino',layout)
workbook_in = pyXL.load_workbook('Exercicios.xlsx')
exercise_sheets = workbook_in['Exercicios']


while True:
    event, values = window.read()

    if event == sg.WIN_CLOSED:
        break

    if event == '-GENERATE-':
        exercisesSelected = []
        exercisesExcluded = []
        for line in exercise_sheets.iter_rows(min_row=2):
            if line[1].value == values['-MODALITY-']:
                exercisesSelected.append(exercise(line[0].value, line[2].value))
            else:
                exercisesExcluded.append(exercise(line[0].value, line[2].value))

        if values['-MODALITY-'] == 'Both':
            exercisesSelected = exercisesExcluded

        workbook_out = pyXL.load_workbook('PlanejamentoDeTreino.xlsx')
        exit_sheet = workbook_out.active
        
        for iten in exercisesSelected:
            print(iten)

        i = 0 
        while i in range(len(exercisesSelected)): 
            exit_sheet[f'A{i+1}'].value = exercisesSelected[i].getName()
            i += 1
        
        workbook_out.save('PlanejamentoDeTreino.xlsx')


window.close()